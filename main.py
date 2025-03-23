import os
from github import Github
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import load_workbook
from dotenv import load_dotenv

# Retrieve token from environment variable
# Authentication is defined via GitHub.Auth
from github import Auth

load_dotenv()
# Using an access token
auth = Auth.Token(os.getenv("GITHUB_TOKEN"))

# Authenticate with GitHub
g = Github(auth=auth)

# 📅 Define the time range for retrieving data (last 13 days)
today = datetime.now()
start_date = today - timedelta(days=13)  # Only the last 13 days
date_strs = [d.strftime("%Y-%m-%d") for d in pd.date_range(start=start_date, end=today, freq="D")]

# 📂 Attempt to open the Excel file (preserving design)
file_path = "Book1.xlsx"
try:
    wb = load_workbook(file_path)
    ws = wb.active
    df = pd.read_excel(file_path, index_col=0, dtype=str).fillna(0).astype(np.int16)
except FileNotFoundError:
    raise FileNotFoundError("⚠ File 'Book1.xlsx' not found! Please create the file first.")

# ✨ Check and add new dates without changing the style
for date in date_strs:
    if date not in df.columns:
        df[date] = 0

# 📊 Retrieve clone data from GitHub
# Loop checks each repo one by one
for repo in g.get_user().get_repos():
    if repo.private:
        continue

    print(f"🔍 Retrieving data for repository: {repo.name}...")

    try:
        clone_data = repo.get_clones_traffic(per="day")
        if not clone_data or "clones" not in clone_data.raw_data:
            print(f"⚠ Clone data for {repo.name} is not available.")
            continue

        # Create object of clone count {"date":count, "date":count, ....}
        clone_count = {str(data["timestamp"])[:10]: data["count"] for data in clone_data.raw_data["clones"]}

        if repo.name in df.index:
            # Update the existing repository data
            for date in date_strs:
                if date in clone_count:
                    df.at[repo.name, date] = clone_count[date]
        else:
            print(f"⚠ New repository ({repo.name}) detected, updating data...")

            # Find the first empty row in the Excel sheet for the new repository
            repo_row = None
            for row_idx in range(2, ws.max_row + 1):
                if ws.cell(row=row_idx, column=1).value is None:
                    repo_row = row_idx
                    break

            # If no empty row found, add a new row at the end
            if not repo_row:
                repo_row = ws.max_row + 1

            # Insert repository name in the first column of the empty row
            ws.cell(row=repo_row, column=1, value=repo.name)
            # Insert the clone data for the last 13 days
            for col_idx, date in enumerate(df.columns, start=2):
                if date in clone_count:
                    ws.cell(row=repo_row, column=col_idx, value=clone_count[date])

    except Exception as e:
        print(f"❌ Error retrieving data for {repo.name}: {e}")

# 📌 **Update column headers (date titles)**
for col_idx, date in enumerate(df.columns, start=2):
    ws.cell(row=1, column=col_idx, value=date)

# 📌 **Update values in Excel without adding new rows**
for repo_name in df.index:
    # Check if repository already exists in the Excel file
    repo_row = None
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == repo_name:
            repo_row = row_idx
            break

    # If repository is found, update values
    if repo_row:
        for col_idx, date in enumerate(df.columns, start=2):
            ws.cell(row=repo_row, column=col_idx, value=df.at[repo_name, date])
    else:
        # If repository is new, find the first empty row and insert
        for row_idx in range(2, ws.max_row + 1):
            if ws.cell(row=row_idx, column=1).value is None:
                repo_row = row_idx
                break

        # If no empty row found, add a new row at the end
        if not repo_row:
            repo_row = ws.max_row + 1

        # Insert repository data in the first empty row
        ws.cell(row=repo_row, column=1, value=repo_name)
        for col_idx, date in enumerate(df.columns, start=2):
            ws.cell(row=repo_row, column=col_idx, value=df.at[repo_name, date])

# 💾 Save changes without adding `Unnamed` columns
wb.save(file_path)
g.close()

print("✅ Data has been updated!")
